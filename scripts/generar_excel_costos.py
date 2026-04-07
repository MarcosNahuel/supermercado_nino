"""Genera Excel de Estructura de Costos - Rotiseria NINO
Con formulas Excel vinculadas entre celdas."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings("ignore")

wb = openpyxl.Workbook()

# === ESTILOS ===
titulo_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
sub_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
norm = Font(name="Calibri", size=10)
bold = Font(name="Calibri", size=10, bold=True)
note_font = Font(name="Calibri", size=9, italic=True, color="996600")
red_b = Font(name="Calibri", size=10, bold=True, color="CC0000")
grn_b = Font(name="Calibri", size=10, bold=True, color="006600")
det_font = Font(name="Calibri", size=9, color="666666")

tit_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
lblue = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
lgrn = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
lyel = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
lred = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
lgray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")

brd = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def write_header(ws, row, vals, c0=1):
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=c0 + i, value=v)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = brd


def wr(ws, row, vals, c0=1, ft=norm, fl=None):
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=c0 + i, value=v)
        c.font = ft
        c.border = brd
        if fl:
            c.fill = fl
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.alignment = Alignment(horizontal="right")
            if isinstance(v, float):
                c.number_format = "#,##0"
            elif abs(v) >= 100:
                c.number_format = "#,##0"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def titulo_row(ws, row, text, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = titulo_font
    c.fill = tit_fill
    c.alignment = Alignment(horizontal="center")


def seccion_row(ws, row, text, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = sub_font
    c.fill = lblue


def nota_row(ws, row, text, cols=5):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=text).font = note_font


def fml(ws, row, col, formula, fmt="#,##0"):
    """Sobreescribe celda con formula Excel, preservando formato visual."""
    c = ws.cell(row=row, column=col)
    c.value = formula
    c.number_format = fmt
    c.alignment = Alignment(horizontal="right")


# === DATOS ===
df = pd.read_parquet("data/processed/detalle_lineas.parquet")
roti = df[df["categoria"] == "ROTISERIA"].copy()
roti["fecha"] = pd.to_datetime(roti["fecha"])

rc = roti[roti["mes_incompleto"] == False] if "mes_incompleto" in roti.columns else roti
rc["fecha"] = pd.to_datetime(rc["fecha"])
vm = rc.groupby(rc["fecha"].dt.to_period("M")).agg(
    ventas=("importe_total", "sum"),
    tickets=("ticket_id", "nunique"),
    cantidad=("cantidad", "sum"),
)

TK = round(vm["tickets"].mean())

# --- Indice estacional 2025 (meses completos) ---
roti_2025 = roti[roti["fecha"].dt.year == 2025].copy()
if "mes_incompleto" in roti_2025.columns:
    roti_2025 = roti_2025[roti_2025["mes_incompleto"] == False]
monthly_2025 = roti_2025.groupby(roti_2025["fecha"].dt.to_period("M"))["importe_total"].sum()
daily_rates_2025 = {}
for m, v in monthly_2025.items():
    days_in_month = m.end_time.day
    daily_rates_2025[m.month] = v / days_in_month
avg_daily_2025 = np.mean(list(daily_rates_2025.values()))
seasonal_index = {m: dr / avg_daily_2025 for m, dr in daily_rates_2025.items()}

# --- Ventas reales ultimos 31 dias (2026) desestacionalizadas ---
EXCEL_DIR = "H:/Mi unidad/PYMEINSIDE/nino/extraccion datos/2026/marzo"
excel_files = [
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes1.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes2.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes22.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes3.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes18-28.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes31.xlsx",
]
all_rotis_2026 = []
for fp in excel_files:
    df_ex = pd.read_excel(fp, header=3)
    if "Departamento" in df_ex.columns:
        mask = df_ex["Departamento"].str.contains("ROTIS", case=False, na=False)
        all_rotis_2026.append(df_ex[mask])
df_2026 = pd.concat(all_rotis_2026, ignore_index=True)
df_2026["Fecha"] = pd.to_datetime(df_2026["Fecha"], format="%d/%m/%y", errors="coerce")
df_2026 = df_2026.drop_duplicates(subset=["Fecha", "Comprobante", "Código"])
df_2026 = df_2026[df_2026["Fecha"].notna()]

date_max_2026 = df_2026["Fecha"].max()
cutoff_2026 = date_max_2026 - pd.Timedelta(days=30)
df_window = df_2026[df_2026["Fecha"] >= cutoff_2026]
WINDOW_DAYS = df_window["Fecha"].dt.date.nunique()
ventas_window = df_window["Importe"].sum()

days_per_month = df_window.groupby(df_window["Fecha"].dt.month)["Fecha"].apply(
    lambda x: x.dt.date.nunique()
)
idx_weighted = sum(
    seasonal_index.get(m, 1.0) * d for m, d in days_per_month.items()
) / WINDOW_DAYS

V = round((ventas_window / WINDOW_DAYS) * 30.4167 / idx_weighted)
WINDOW_START = cutoff_2026.strftime("%d/%m/%Y")
WINDOW_END = date_max_2026.strftime("%d/%m/%Y")

mp = roti.groupby("tipo_medio_pago")["importe_total"].sum()
mp_p = mp / mp.sum()
pEF = mp_p.get("EFECTIVO", 0)
pDE = mp_p.get("TARJETA DE DEBITO", mp_p.get("TARJETA DE DÉBITO", 0))
pCR = mp_p.get("TARJETA DE CREDITO", mp_p.get("TARJETA DE CRÉDITO", 0))
pBI = sum(v for k, v in mp_p.items() if "BILLETERA" in k.upper())

# Parametros
SB_REM = 1_080_274       # basico remunerativo (base contribuciones F.930)
SB_NR = 100_000          # no remunerativo (no genera cargas patronales)
SB_TOTAL = SB_REM + SB_NR
NE = 3
CP_PCT = 0.2974  # 12.71% SIPA + 1.58% PAMI + 5.56% AAFF + 0.89% FNE + 6% OS + 3% ART
MG = 0.30
cCR = 0.03
cDE = 0.012
cBI = 0.015
IIBB_PCT = 0.03   # 3% Ley Impositiva Mendoza 2026
ICH_PCT = 0.012    # 0.6% + 0.6% Ley 25.413 - solo sobre bancarizado
MERMA_PCT = 0.05

# Calculos (Python - para validacion y hojas sin formulas)
CMV = round(V * (1 - MG))
MB = round(V * MG)
ST_total = SB_TOTAL * NE
ST_rem = SB_REM * NE
CP_m = round(ST_rem * CP_PCT)
SAC = round((ST_rem + CP_m) / 12)
SEG = NE * 500
LAB = ST_total + CP_m + SAC + SEG

# Servicios - Basado en facturas reales Feb-Mar 2026
GAS_FACTURA = 664_000
GAS_FACTURA_M3 = 1_485
GAS_M3_ROTIS = 332      # 22.4% - horno convector 3.5h/dia + cocina 2h/dia
GAS_M3_PANAD = GAS_FACTURA_M3 - GAS_M3_ROTIS  # 1,153 m3 panaderia (77.6%)
GAS_COSTO_M3 = GAS_FACTURA / GAS_FACTURA_M3
gas_m = round(GAS_M3_ROTIS * GAS_COSTO_M3)

ELEC_KWH = 500
ELEC_FACTURA = 3_463_771
ELEC_FACTURA_KWH = 9_470
ELEC_COSTO_KWH = ELEC_FACTURA / ELEC_FACTURA_KWH
elec_m = round(ELEC_KWH * ELEC_COSTO_KWH)

SERV = gas_m + elec_m

# Otros costos fijos (SIN alquiler, amortizacion equipos, agua, seguros)
HAB_MUNI = 8_000
HAB_BROM = 12_000
INDUM = 15_000
MANT = 30_000
DESC = 80_000
OTROS_F = HAB_MUNI + HAB_BROM + INDUM + MANT + DESC

comCR = round(V * pCR * cCR)
comDE = round(V * pDE * cDE)
comBI = round(V * pBI * cBI)
comT = comCR + comDE + comBI
iibb_m = round(V * IIBB_PCT)
pBANC = 1 - pEF
ich_m = round(V * pBANC * ICH_PCT)
merma_m = round(V * MERMA_PCT)
CV = comT + iibb_m + ich_m + merma_m

TF = LAB + SERV + OTROS_F
RES = MB - TF - CV

# =============================================
# HOJA 1: ESTRUCTURA DE COSTOS
# =============================================
ws = wb.active
ws.title = "Estructura Costos"
set_widths(ws, [48, 20, 14, 14, 42])
R = {}  # Tracker de filas para formulas

titulo_row(ws, 1, "ESTRUCTURA DE COSTOS - ROTISERIA NINO (BOSIN S.A.)")
nota_row(ws, 2, f"Base: ultimos {WINDOW_DAYS} dias ({WINDOW_START} - {WINDOW_END}) desestacionalizado | Elaborado: Marzo 2026")

r = 4
seccion_row(ws, r, "1. VENTAS")
r = 5
write_header(ws, r, ["Concepto", "Monto Mensual ($)", "% s/Ventas", "Tasa/Base", "Notas"])
r = 6
wr(ws, r, ["VENTAS BRUTAS (con IVA)", round(V * 1.21), "", "121%", "Facturacion total IVA incluido"])
R['v_bruta'] = r
r = 7
wr(ws, r, ["(-) IVA Debito Fiscal", round(-V * 0.21), "", "21%", "Se compensa con credito fiscal de compras"])
R['iva_df'] = r
r = 8
wr(ws, r, ["VENTAS NETAS (sin IVA)", V, "100.0%", "", "Base para todos los calculos"], ft=bold, fl=lgrn)
R['ventas'] = r
r = 9
wr(ws, r, [f"  Tickets promedio/mes: {TK:,} | Ticket prom: ${round(V/TK):,}", "", "", "", ""], ft=det_font)
r = 10
wr(ws, r, [f"  Efectivo {pEF*100:.1f}% | Debito {pDE*100:.1f}% | Credito {pCR*100:.1f}% | Billetera {pBI*100:.1f}%", "", "", "", ""], ft=det_font)

# --- CMV por categoria de producto ---
roti_prods = roti.groupby("descripcion")["importe_total"].sum()
def clasificar(nombre):
    n = nombre.upper()
    if "EMPANADA" in n:
        return "Empanadas"
    elif any(k in n for k in ["MATAMBRE", "PERNIL", "LENGUA", "ARROLLADO", "BONDIOLA", "CERDO", "POLLO RELLENO"]):
        return "Carnes elaboradas"
    elif any(k in n for k in ["PIZZA", "CANELONE"]):
        return "Pizzas y canelones"
    elif any(k in n for k in ["MIGA", "MILANES"]):
        return "Milanesas"
    elif "SALSA" in n:
        return "Salsas"
    else:
        return "Otros elaborados"

familia_ventas = {}
for prod, venta in roti_prods.items():
    fam = clasificar(prod)
    familia_ventas[fam] = familia_ventas.get(fam, 0) + venta
total_hist = sum(familia_ventas.values())

familias_orden = ["Empanadas", "Carnes elaboradas", "Pizzas y canelones",
                  "Milanesas", "Salsas", "Otros elaborados"]
insumos_principales = {
    "Empanadas": "Carne, masa, queso, verdura, aceite",
    "Carnes elaboradas": "Matambre, pernil, lengua, pollo, condimentos",
    "Pizzas y canelones": "Harina, muzzarella, salsa, carne, verdura",
    "Milanesas": "Carne, pan rallado, huevo, aceite",
    "Salsas": "Tomate, carne, condimentos",
    "Otros elaborados": "Varios (tartas, fiambres, preparaciones)",
}

r = 12
seccion_row(ws, r, "2. COSTO DE MERCADERIA VENDIDA (CMV)")
r += 1
write_header(ws, r, ["Concepto", "Monto Mensual ($)", "% s/Ventas", "% del CMV", "Insumos principales"])
r += 1
wr(ws, r, ["COSTO TOTAL INSUMOS", -CMV, "70.0%", "100%", "Margen bruto 30% s/PVP (dato del sistema)"], ft=bold, fl=lred)
R['cmv'] = r
for fam in familias_orden:
    v_fam = familia_ventas.get(fam, 0)
    pct_mix = v_fam / total_hist
    cmv_fam = round(V * pct_mix * (1 - MG))
    r += 1
    wr(ws, r, [f"  {fam}", -cmv_fam, f"{cmv_fam/V*100:.1f}%",
               f"{pct_mix*100:.0f}%", insumos_principales.get(fam, "")], ft=det_font)
r += 1
wr(ws, r, ["MARGEN BRUTO", MB, "30.0%", "", "Ventas - CMV"], ft=bold, fl=lgrn)
R['mb'] = r
r += 1
nota_row(ws, r, "  Sensibilizar: 25%, 30%, 35% (ver hoja Sensibilizacion)")

# --- COSTOS FIJOS ---
r += 2
seccion_row(ws, r, "3. COSTOS FIJOS")
r += 1
write_header(ws, r, ["Concepto", "Monto Mensual ($)", "% s/Ventas", "Tasa/Base", "Notas"])

r += 1
wr(ws, r, ["3.1 SUELDOS Y CARGAS SOCIALES", "", "", "", ""], ft=bold, fl=lgray)
r += 1
wr(ws, r, [f"Sueldos brutos ({NE} empleados)", -ST_total, f"{ST_total/V*100:.1f}%", f"${SB_TOTAL:,} c/u", "CCT 130/75 Aux.Esp.A (Mar 2026)"])
R['sueldos'] = r
r += 1
wr(ws, r, ["  - Remunerativo", -ST_rem, "", f"${SB_REM:,} c/u", "Base contribuciones (F.930)"], ft=det_font)
R['sueldos_rem'] = r
r += 1
wr(ws, r, ["  - No Remunerativo", -(SB_NR * NE), "", f"${SB_NR:,} c/u", "No genera cargas patronales"], ft=det_font)
R['sueldos_nr'] = r
r += 1
wr(ws, r, ["Contribuciones patronales (F.930)", -CP_m, f"{CP_m/V*100:.1f}%", "29.74% s/rem", f"Base: ${ST_rem:,} (solo remunerativo)"])
R['contrib'] = r
r += 1
wr(ws, r, ["  - Jubilacion (SIPA)", round(-ST_rem * 0.1271), "", "12.71%", "Alicuota NO MiPyME servicios/comercio"], ft=det_font)
R['sipa'] = r
r += 1
wr(ws, r, ["  - PAMI (Ley 19.032)", round(-ST_rem * 0.0158), "", "1.58%", ""], ft=det_font)
R['pami'] = r
r += 1
wr(ws, r, ["  - Asignaciones Familiares", round(-ST_rem * 0.0556), "", "5.56%", ""], ft=det_font)
R['aaff'] = r
r += 1
wr(ws, r, ["  - Fondo Nacional de Empleo", round(-ST_rem * 0.0089), "", "0.89%", ""], ft=det_font)
R['fne'] = r
r += 1
wr(ws, r, ["  - Obra Social patronal", round(-ST_rem * 0.06), "", "6.00%", ""], ft=det_font)
R['os'] = r
r += 1
wr(ws, r, ["  - ART (estimado)", round(-ST_rem * 0.03), "", "~3.00%", "Varia segun siniestralidad"], ft=det_font)
R['art'] = r
r += 1
wr(ws, r, ["SAC proporcional (1/12)", -SAC, f"{SAC/V*100:.1f}%", "", "Aguinaldo s/remunerativo + cargas"])
R['sac'] = r
r += 1
wr(ws, r, ["Seguro Vida Obligatorio", -SEG, "", f"~$500 c/u", "Decreto 1567/74"], ft=det_font)
R['seg'] = r
r += 1
wr(ws, r, ["SUBTOTAL LABORAL", -LAB, f"{LAB/V*100:.1f}%", "", ""], ft=bold, fl=lyel)
R['sub_lab'] = r

r += 2
wr(ws, r, ["3.2 SERVICIOS (prorrateo rotiseria)", "", "", "", ""], ft=bold, fl=lgray)
r += 1
wr(ws, r, [f"Gas natural ({GAS_M3_ROTIS} m3/mes)", -gas_m, f"{gas_m/V*100:.1f}%", f"${GAS_COSTO_M3:.0f}/m3", "Prorrateo s/factura Ecogas Patricios 344 Feb 2026"])
R['gas'] = r
r += 1
wr(ws, r, [f"  Factura: ${GAS_FACTURA:,}/{GAS_FACTURA_M3}m3. Rotis {GAS_M3_ROTIS}m3 ({GAS_M3_ROTIS/GAS_FACTURA_M3*100:.0f}%) + Panad {GAS_M3_PANAD}m3 ({GAS_M3_PANAD/GAS_FACTURA_M3*100:.0f}%) = 100%", "", "", "", ""], ft=det_font)
r += 1
wr(ws, r, ["  Equipamiento: horno convector 8 band. ~4.3h/dia + cocina 2 horn. ~2h/dia", "", "", "", ""], ft=det_font)
r += 1
wr(ws, r, [f"Electricidad ({ELEC_KWH} kWh/mes)", -elec_m, f"{elec_m/V*100:.1f}%", f"${ELEC_COSTO_KWH:.0f}/kWh", "Prorrateo s/factura EDEMSA Patricios 334 Mar 2026"])
R['elec'] = r
r += 1
wr(ws, r, [f"  Factura real: ${ELEC_FACTURA:,} / {ELEC_FACTURA_KWH} kWh = ${ELEC_COSTO_KWH:.0f}/kWh all-in (IVA 27% + imp.)", "", "", "", ""], ft=det_font)
r += 1
wr(ws, r, ["  Equipamiento: vitrina exhibidora + heladera insumos + campana + iluminacion", "", "", "", ""], ft=det_font)
r += 1
wr(ws, r, ["SUBTOTAL SERVICIOS", -SERV, f"{SERV/V*100:.1f}%", "", "Basado en facturas reales Feb-Mar 2026"], ft=bold, fl=lyel)
R['sub_serv'] = r

r += 2
wr(ws, r, ["3.3 OTROS COSTOS FIJOS (estimados)", "", "", "", ""], ft=bold, fl=lgray)
otros_items = [
    ("Descartables / packaging", -DESC, "Bandejas aluminio, bolsas, film, envases"),
    ("Mantenimiento equipos", -MANT, "Service horno + heladera, repuestos amort."),
    ("Indumentaria / EPP", -INDUM, "Delantales, gorros, guantes (3 empleados)"),
    ("Habilitacion bromatologica", -HAB_BROM, "Anual / 12. Obligatoria elaboracion alimentos"),
    ("Habilitacion municipal", -HAB_MUNI, "Tasa comercio Lujan de Cuyo / 12"),
]
R['otros_start'] = r + 1
for item, monto, nota in otros_items:
    r += 1
    wr(ws, r, [f"  {item}", monto, f"{abs(monto)/V*100:.1f}%", "", nota], fl=lyel)
R['otros_end'] = r
r += 1
wr(ws, r, ["SUBTOTAL OTROS FIJOS", -OTROS_F, f"{OTROS_F/V*100:.1f}%", "", "Estimados - ajustar con datos reales"], ft=bold, fl=lyel)
R['sub_otros'] = r

r += 2
wr(ws, r, ["TOTAL COSTOS FIJOS", -TF, f"{TF/V*100:.1f}%", "", "Laboral + Servicios + Otros"], ft=bold, fl=lred)
R['total_cf'] = r

# --- COSTOS VARIABLES ---
r += 2
seccion_row(ws, r, "4. COSTOS VARIABLES (% sobre ventas)")
r += 1
write_header(ws, r, ["Concepto", "Monto Mensual ($)", "% s/Ventas", "Tasa/Base", "Notas"])

r += 1
wr(ws, r, ["4.1 COMISIONES TARJETAS", "", "", "", ""], ft=bold, fl=lgray)
r += 1
wr(ws, r, [f"Tarjeta Credito ({pCR*100:.1f}% ventas)", -comCR, f"{comCR/V*100:.2f}%", "3.0%", "Naranja 41%, Visa 41%, MC 15%"])
R['comCR'] = r
r += 1
wr(ws, r, [f"Tarjeta Debito ({pDE*100:.1f}% ventas)", -comDE, f"{comDE/V*100:.2f}%", "1.2%", "Visa 64%, Mastercard 36%"])
R['comDE'] = r
r += 1
wr(ws, r, [f"Billetera Virtual ({pBI*100:.1f}% ventas)", -comBI, f"{comBI/V*100:.2f}%", "1.5%", "Mercado Pago y otras"])
R['comBI'] = r
r += 1
wr(ws, r, ["Subtotal comisiones", -comT, f"{comT/V*100:.2f}%", "", ""], ft=bold)
R['sub_com'] = r

r += 2
wr(ws, r, ["4.2 IMPUESTOS OPERATIVOS", "", "", "", ""], ft=bold, fl=lgray)
r += 1
wr(ws, r, ["Ingresos Brutos (Mendoza)", -iibb_m, f"{IIBB_PCT*100:.1f}%", f"{IIBB_PCT*100:.0f}%", "Ley Impositiva Mza 2026, Rubro 7 comercio minorista"])
R['iibb'] = r
r += 1
wr(ws, r, ["Imp. Debitos y Creditos Bancarios", -ich_m, f"{ich_m/V*100:.2f}%", "1.2% s/banc.", f"Ley 25.413. Solo s/bancarizado ({pBANC*100:.1f}% ventas, excluye efectivo)"])
R['ich'] = r
r += 1
wr(ws, r, ["Subtotal impuestos", -(iibb_m + ich_m), f"{(iibb_m+ich_m)/V*100:.1f}%", "", ""], ft=bold)
R['sub_imp'] = r

r += 2
wr(ws, r, ["4.3 MERMA / DESCARTE", "", "", "", ""], ft=bold, fl=lgray)
r += 1
wr(ws, r, ["Merma estimada", -merma_m, f"{MERMA_PCT*100:.1f}%", "", "Rotiseria tipico 3-8%. AJUSTAR con dato real"], fl=lyel)
R['merma'] = r

r += 2
wr(ws, r, ["TOTAL COSTOS VARIABLES", -CV, f"{CV/V*100:.1f}%", "", ""], ft=bold, fl=lred)
R['total_cv'] = r

# --- RESULTADO ---
r += 2
seccion_row(ws, r, "5. RESULTADO - MARGEN DE CONTRIBUCION")
r += 1
write_header(ws, r, ["Concepto", "Monto Mensual ($)", "% s/Ventas", "", "Notas"])
r += 1
wr(ws, r, ["Ventas Netas", V, "100.0%", "", ""])
R['res_ventas'] = r
r += 1
wr(ws, r, ["(-) CMV", -CMV, "70.0%", "", ""])
R['res_cmv'] = r
r += 1
wr(ws, r, ["MARGEN BRUTO", MB, "30.0%", "", ""], ft=bold)
R['res_mb'] = r
r += 1
wr(ws, r, ["(-) Costos Fijos", -TF, f"{TF/V*100:.1f}%", "", "Laboral + Servicios + Otros"])
R['res_cf'] = r
r += 1
wr(ws, r, ["(-) Costos Variables", -CV, f"{CV/V*100:.1f}%", "", ""])
R['res_cv'] = r
r += 1
estado = "SUPERAVITARIO" if RES > 0 else "DEFICITARIO"
ft_r = grn_b if RES > 0 else red_b
fl_r = lgrn if RES > 0 else lred
wr(ws, r, ["RESULTADO NETO", RES, f"{RES/V*100:.1f}%", "", estado], ft=ft_r, fl=fl_r)
R['resultado'] = r

r += 1
nota_row(ws, r, "  Items en amarillo son estimados. Ajustar con datos reales.")

# --- PUNTO DE EQUILIBRIO ---
r += 2
seccion_row(ws, r, "6. PUNTO DE EQUILIBRIO")
var_pct = CV / V
veq = round(TF / (MG - var_pct))
r += 1
wr(ws, r, [f"Ventas necesarias para cubrir costos fijos (margen {MG*100:.0f}%)", veq, "", "", ""], ft=bold)
R['veq'] = r
r += 1
wr(ws, r, ["Ventas actuales promedio", V, "", "", ""])
R['v_actual'] = r
cob = V / veq * 100
r += 1
fl_cob = lgrn if cob > 100 else lred
est_cob = "OK - cubre costos fijos" if cob > 100 else "INSUFICIENTE"
wr(ws, r, ["Cobertura punto de equilibrio", "", f"{cob:.0f}%", "", est_cob], ft=bold, fl=fl_cob)
R['cobertura'] = r

# --- SUPUESTOS ---
r += 2
seccion_row(ws, r, "SUPUESTOS")
supuestos = [
    "Empresa NO inscripta como MiPyME - Contribuciones patronales SUSS al 20.74%",
    "CCT 130/75 Empleados de Comercio - Categoria Auxiliar Especializado A, sin antiguedad",
    f"Sueldo: ${SB_REM:,} remunerativo + ${SB_NR:,} NR = ${SB_TOTAL:,}/empleado (escala Dic 2025 - Mar 2026)",
    "Contribuciones patronales (29.74%) calculadas SOLO sobre remunerativo. NR no genera cargas",
    "SAC mensualizado sobre base remunerativa + contribuciones",
    "Margen bruto sobre PVP: 30% (parametro del sistema de facturacion)",
    "Comisiones: Credito 3%, Debito 1.2%, Billetera 1.5% (promedios de mercado)",
    f"IIBB Mendoza: {IIBB_PCT*100:.0f}% (Ley Impositiva Mza 2026, Rubro 7 comercio minorista)",
    f"Imp. Debitos/Creditos: 1.2% SOLO sobre bancarizado ({pBANC*100:.1f}% ventas). Ley 25.413",
    "Merma estimada: 5% (ajustar con medicion real, tipico rotiseria 3-8%)",
    "IVA: no se incluye como costo (se compensa DF vs CF). Solo se muestra para referencia",
    f"Ventas: ultimos {WINDOW_DAYS} dias reales ({WINDOW_START} - {WINDOW_END}) desestacionalizados con indice 2025",
    f"Gas: {GAS_M3_ROTIS} m3/mes (horno convector + cocina). Factura ${GAS_FACTURA:,}/{GAS_FACTURA_M3}m3 compartida con panaderia",
    f"Electricidad: {ELEC_KWH} kWh/mes. Tarifa ${ELEC_COSTO_KWH:.2f}/kWh all-in s/factura EDEMSA",
    "NO incluye: alquiler, amortizacion equipos, agua, seguros (excluidos por decision del cliente)",
]
for i, s in enumerate(supuestos):
    r += 1
    nota_row(ws, r, f"  {i+1}. {s}")


# =============================================
# FORMULAS EXCEL - Hoja 1
# =============================================
B = 2   # columna Monto Mensual
C = 3   # columna % s/Ventas
nf = "#,##0"
pf = "0.0%"

# --- Seccion 1: Ventas ---
fml(ws, R['v_bruta'], B, f"=B{R['ventas']}*1.21")
fml(ws, R['iva_df'], B, f"=-B{R['ventas']}*0.21")

# --- Seccion 2: CMV ---
fml(ws, R['cmv'], B, f"=-B{R['ventas']}*{1-MG}")
fml(ws, R['cmv'], C, f"=ABS(B{R['cmv']})/B{R['ventas']}", pf)
fml(ws, R['mb'], B, f"=B{R['ventas']}+B{R['cmv']}")
fml(ws, R['mb'], C, f"=B{R['mb']}/B{R['ventas']}", pf)

# --- Seccion 3: Costos Fijos ---
# Sueldos = Remunerativo + NR
fml(ws, R['sueldos'], B, f"=B{R['sueldos_rem']}+B{R['sueldos_nr']}")
fml(ws, R['sueldos'], C, f"=ABS(B{R['sueldos']})/B{R['ventas']}", pf)
# Contribuciones patronales
fml(ws, R['contrib'], B, f"=B{R['sueldos_rem']}*{CP_PCT}")
fml(ws, R['contrib'], C, f"=ABS(B{R['contrib']})/B{R['ventas']}", pf)
# Desglose contribuciones
fml(ws, R['sipa'], B, f"=B{R['sueldos_rem']}*0.1271")
fml(ws, R['pami'], B, f"=B{R['sueldos_rem']}*0.0158")
fml(ws, R['aaff'], B, f"=B{R['sueldos_rem']}*0.0556")
fml(ws, R['fne'], B, f"=B{R['sueldos_rem']}*0.0089")
fml(ws, R['os'], B, f"=B{R['sueldos_rem']}*0.06")
fml(ws, R['art'], B, f"=B{R['sueldos_rem']}*0.03")
# SAC = (remunerativo + contribuciones) / 12
fml(ws, R['sac'], B, f"=(B{R['sueldos_rem']}+B{R['contrib']})/12")
fml(ws, R['sac'], C, f"=ABS(B{R['sac']})/B{R['ventas']}", pf)
# Subtotal laboral
fml(ws, R['sub_lab'], B, f"=B{R['sueldos']}+B{R['contrib']}+B{R['sac']}+B{R['seg']}")
fml(ws, R['sub_lab'], C, f"=ABS(B{R['sub_lab']})/B{R['ventas']}", pf)

# Servicios
fml(ws, R['sub_serv'], B, f"=B{R['gas']}+B{R['elec']}")
fml(ws, R['sub_serv'], C, f"=ABS(B{R['sub_serv']})/B{R['ventas']}", pf)

# Otros fijos
fml(ws, R['sub_otros'], B, f"=SUM(B{R['otros_start']}:B{R['otros_end']})")
fml(ws, R['sub_otros'], C, f"=ABS(B{R['sub_otros']})/B{R['ventas']}", pf)

# Total costos fijos
fml(ws, R['total_cf'], B, f"=B{R['sub_lab']}+B{R['sub_serv']}+B{R['sub_otros']}")
fml(ws, R['total_cf'], C, f"=ABS(B{R['total_cf']})/B{R['ventas']}", pf)

# --- Seccion 3: % faltantes Costos Fijos ---
fml(ws, R['gas'], C, f"=ABS(B{R['gas']})/B{R['ventas']}", pf)
fml(ws, R['elec'], C, f"=ABS(B{R['elec']})/B{R['ventas']}", pf)
# % otros fijos individuales
for row_i in range(R['otros_start'], R['otros_end'] + 1):
    fml(ws, row_i, C, f"=ABS(B{row_i})/B{R['ventas']}", pf)

# --- Seccion 4: Costos Variables ---
# Comisiones vinculadas a ventas
fml(ws, R['comCR'], B, f"=-B{R['ventas']}*{pCR:.6f}*{cCR}")
fml(ws, R['comCR'], C, f"=ABS(B{R['comCR']})/B{R['ventas']}", "0.00%")
fml(ws, R['comDE'], B, f"=-B{R['ventas']}*{pDE:.6f}*{cDE}")
fml(ws, R['comDE'], C, f"=ABS(B{R['comDE']})/B{R['ventas']}", "0.00%")
fml(ws, R['comBI'], B, f"=-B{R['ventas']}*{pBI:.6f}*{cBI}")
fml(ws, R['comBI'], C, f"=ABS(B{R['comBI']})/B{R['ventas']}", "0.00%")
fml(ws, R['sub_com'], B, f"=B{R['comCR']}+B{R['comDE']}+B{R['comBI']}")
fml(ws, R['sub_com'], C, f"=ABS(B{R['sub_com']})/B{R['ventas']}", "0.00%")
# Impuestos
fml(ws, R['iibb'], B, f"=-B{R['ventas']}*{IIBB_PCT}")
fml(ws, R['iibb'], C, f"=ABS(B{R['iibb']})/B{R['ventas']}", pf)
fml(ws, R['ich'], B, f"=-B{R['ventas']}*{pBANC:.6f}*{ICH_PCT}")
fml(ws, R['ich'], C, f"=ABS(B{R['ich']})/B{R['ventas']}", "0.00%")
fml(ws, R['sub_imp'], B, f"=B{R['iibb']}+B{R['ich']}")
fml(ws, R['sub_imp'], C, f"=ABS(B{R['sub_imp']})/B{R['ventas']}", pf)
# Merma
fml(ws, R['merma'], B, f"=-B{R['ventas']}*{MERMA_PCT}")
fml(ws, R['merma'], C, f"=ABS(B{R['merma']})/B{R['ventas']}", pf)
# Total CV
fml(ws, R['total_cv'], B, f"=B{R['sub_com']}+B{R['sub_imp']}+B{R['merma']}")
fml(ws, R['total_cv'], C, f"=ABS(B{R['total_cv']})/B{R['ventas']}", pf)

# --- Seccion 5: Resultado ---
fml(ws, R['res_ventas'], B, f"=B{R['ventas']}")
fml(ws, R['res_cmv'], B, f"=B{R['cmv']}")
fml(ws, R['res_cmv'], C, f"=ABS(B{R['res_cmv']})/B{R['res_ventas']}", pf)
fml(ws, R['res_mb'], B, f"=B{R['res_ventas']}+B{R['res_cmv']}")
fml(ws, R['res_mb'], C, f"=B{R['res_mb']}/B{R['res_ventas']}", pf)
fml(ws, R['res_cf'], B, f"=B{R['total_cf']}")
fml(ws, R['res_cf'], C, f"=ABS(B{R['res_cf']})/B{R['res_ventas']}", pf)
fml(ws, R['res_cv'], B, f"=B{R['total_cv']}")
fml(ws, R['res_cv'], C, f"=ABS(B{R['res_cv']})/B{R['res_ventas']}", pf)
fml(ws, R['resultado'], B, f"=B{R['res_mb']}+B{R['res_cf']}+B{R['res_cv']}")
fml(ws, R['resultado'], C, f"=B{R['resultado']}/B{R['res_ventas']}", pf)

# --- Seccion 6: Punto de equilibrio ---
fml(ws, R['veq'], B, f"=-B{R['total_cf']}/({MG}+B{R['total_cv']}/B{R['ventas']})")
fml(ws, R['v_actual'], B, f"=B{R['ventas']}")
fml(ws, R['cobertura'], C, f"=B{R['v_actual']}/B{R['veq']}", pf)


# =============================================
# HOJA 2: VENTAS MENSUALES
# =============================================
ws2 = wb.create_sheet("Ventas Mensuales")
set_widths(ws2, [14, 18, 10, 10, 14, 16, 16, 16, 16])
titulo_row(ws2, 1, "VENTAS MENSUALES ROTISERIA - DETALLE", 9)
write_header(ws2, 3, ["Periodo", "Ventas ($)", "Tickets", "Items", "Ticket Prom", "Efectivo ($)", "Debito ($)", "Credito ($)", "Billetera ($)"])

roti["mes"] = roti["fecha"].dt.to_period("M")
meses = roti.groupby("mes").agg(
    ventas=("importe_total", "sum"),
    tickets=("ticket_id", "nunique"),
    cantidad=("cantidad", "sum"),
).sort_index()
mp_m = roti.groupby(["mes", "tipo_medio_pago"])["importe_total"].sum().unstack(fill_value=0)

row = 4
for m, d in meses.iterrows():
    v = d["ventas"]
    t = d["tickets"]
    tp = round(v / t) if t > 0 else 0
    ef = mp_m.loc[m].get("EFECTIVO", 0) if m in mp_m.index else 0
    de = 0
    cr = 0
    bi = 0
    if m in mp_m.index:
        for col in mp_m.columns:
            if "DEBIT" in col.upper() or "DÉBITO" in col.upper():
                de = mp_m.loc[m][col]
            elif "CREDIT" in col.upper() or "CRÉDITO" in col.upper():
                cr = mp_m.loc[m][col]
            elif "BILLETERA" in col.upper():
                bi += mp_m.loc[m][col]
    wr(ws2, row, [str(m), round(v), int(t), int(d["cantidad"]), tp, round(ef), round(de), round(cr), round(bi)])
    row += 1

row += 1
wr(ws2, row, [
    "TOTAL", round(meses["ventas"].sum()), int(meses["tickets"].sum()),
    int(meses["cantidad"].sum()), "", "", "", "", ""
], ft=bold, fl=lgrn)


# =============================================
# HOJA 3: PRODUCTOS
# =============================================
ws3 = wb.create_sheet("Productos")
set_widths(ws3, [45, 18, 12, 14, 10, 10])
titulo_row(ws3, 1, "PRODUCTOS ROTISERIA - CLASIFICACION ABC", 6)
write_header(ws3, 3, ["Producto", "Ventas ($)", "Cantidad", "Precio Prom ($)", "% Ventas", "% Acum"])

prods = roti.groupby("descripcion").agg(
    ventas=("importe_total", "sum"),
    cantidad=("cantidad", "sum"),
    pp=("precio_unitario", "mean"),
).sort_values("ventas", ascending=False)
prods["pct"] = prods["ventas"] / prods["ventas"].sum() * 100
prods["acum"] = prods["pct"].cumsum()

row = 4
for p, d in prods.iterrows():
    fl = lgrn if d["acum"] <= 80 else (lyel if d["acum"] <= 95 else None)
    wr(ws3, row, [
        p, round(d["ventas"]), round(d["cantidad"]),
        round(d["pp"]), round(d["pct"], 1), round(d["acum"], 1)
    ], fl=fl)
    row += 1
row += 1
nota_row(ws3, row, "  Verde = Core (80%) | Amarillo = Complementario (80-95%) | Sin color = Marginal", 6)


# =============================================
# HOJA 4: COSTEO POR RECETA
# =============================================
ws_rec = wb.create_sheet("Costeo Recetas")
set_widths(ws_rec, [30, 14, 10, 14, 14, 14, 10, 30])
titulo_row(ws_rec, 1, "COSTEO POR RECETA - INSUMOS ROTISERIA NINO", 8)
nota_row(ws_rec, 2, "Precios insumos: sistema POS supermercado NINO x 0.70 (costo mayorista estimado)", 8)

COST_FACTOR = 0.70

recetas = [
    {
        "nombre": "EMPANADA CARNE x12",
        "pvp": 8_000,
        "pct_ventas": 49.8,
        "items": [
            ("Carne picada especial", 0.35, "kg", 12_200),
            ("Tapas empanada x12", 1, "doc", 1_300),
            ("Cebolla", 0.15, "kg", 1_500),
            ("Huevo (1/2)", 0.5, "u", 292),
            ("Condimentos/grasa", 1, "lote", 300),
        ]
    },
    {
        "nombre": "EMPANADA J&Q x6",
        "pvp": 4_058,
        "pct_ventas": 5.2,
        "items": [
            ("Jamon cocido", 0.12, "kg", 18_525),
            ("Muzzarella", 0.10, "kg", 7_300),
            ("Tapas empanada x6", 0.5, "doc", 1_300),
            ("Condimentos", 1, "lote", 150),
        ]
    },
    {
        "nombre": "MATAMBRE NINO (pieza)",
        "pvp": 38_200,
        "pct_ventas": 13.9,
        "items": [
            ("Matambre vacuno crudo", 1.8, "kg", 18_020),
            ("Condimentos/vinagre", 1, "lote", 500),
        ]
    },
    {
        "nombre": "PIZZA MUZZARELLA",
        "pvp": 6_900,
        "pct_ventas": 2.2,
        "items": [
            ("Harina 0000", 0.25, "kg", 980),
            ("Muzzarella", 0.20, "kg", 7_300),
            ("Salsa tomate", 0.15, "kg", 1_510),
            ("Levadura/sal/aceite", 1, "lote", 200),
        ]
    },
    {
        "nombre": "CANELONES CAR/VER (u)",
        "pvp": 10_329,
        "pct_ventas": 2.2,
        "items": [
            ("Carne picada", 0.15, "kg", 12_200),
            ("Tapas canelones/masa", 1, "u", 800),
            ("Ricota/verdura", 0.15, "kg", 6_550),
            ("Salsa tomate", 0.20, "kg", 1_510),
            ("Muzzarella", 0.08, "kg", 7_300),
        ]
    },
    {
        "nombre": "LENGUA VINAGRETA (pieza)",
        "pvp": 19_000,
        "pct_ventas": 2.9,
        "items": [
            ("Lengua vacuna", 1.5, "kg", 10_150),
            ("Vinagre/condimentos", 1, "lote", 800),
        ]
    },
    {
        "nombre": "PERNIL CERDO (pieza)",
        "pvp": 20_477,
        "pct_ventas": 2.4,
        "items": [
            ("Pernil chancho congelado", 3.0, "kg", 4_500),
            ("Condimentos/adobo", 1, "lote", 800),
        ]
    },
    {
        "nombre": "ARROLLADO POLLO (pieza)",
        "pvp": 22_461,
        "pct_ventas": 2.4,
        "items": [
            ("Suprema pollo", 1.0, "kg", 10_595),
            ("Jamon/queso relleno", 0.15, "kg", 18_525),
            ("Condimentos", 1, "lote", 400),
        ]
    },
    {
        "nombre": "MILANESA MIGA x12",
        "pvp": 8_304,
        "pct_ventas": 2.6,
        "items": [
            ("Nalga/bola lomo", 0.8, "kg", 13_400),
            ("Pan rallado", 0.15, "kg", 3_130),
            ("Huevos x2", 2, "u", 292),
            ("Tapas miga x12", 0.24, "pack50", 15_360),
        ]
    },
    {
        "nombre": "SALSA BOLOGNESA (u)",
        "pvp": 6_850,
        "pct_ventas": 2.2,
        "items": [
            ("Carne picada", 0.30, "kg", 12_200),
            ("Tomate triturado", 0.50, "kg", 1_510),
            ("Cebolla/condimentos", 1, "lote", 400),
        ]
    },
]

# Tabla de precios de insumos
row = 4
seccion_row(ws_rec, row, "PRECIOS DE INSUMOS (sistema POS supermercado)", 8)
row += 1
write_header(ws_rec, row, ["Insumo", "Precio Retail", "Unidad", "Costo Est. (x0.70)", "", "", "", "Fuente"])
insumos_tabla = [
    ("Carne picada especial", 12_200, "$/kg", "CARNICERIA - MOLIDA ESPECIAL"),
    ("Matambre vacuno crudo", 18_020, "$/kg", "CARNICERIA - MATAMBRE"),
    ("Pollo entero", 3_900, "$/kg", "POLLO - AVICOLA LUJAN"),
    ("Suprema pollo", 10_595, "$/kg", "POLLO - SUPREMA"),
    ("Pernil chancho cong.", 4_500, "$/kg", "CONGELADOS - PERNIL CHANCHO"),
    ("Lengua vacuna", 10_150, "$/kg", "CARNICERIA - LENGUA X KG"),
    ("Nalga/bola lomo", 13_400, "$/kg", "CARNICERIA - NALGA/BOLA LOMO"),
    ("Jamon cocido", 18_525, "$/kg", "FIAMBRERIA - PALADINI"),
    ("Muzzarella horma", 7_300, "$/kg", "FIAMBRERIA - JUMIAL HORMA"),
    ("Tapas empanada x12", 1_300, "$/doc", "PANIFIC. - promedio 3 marcas"),
    ("Tapas miga x50", 15_360, "$/pack", "PANIFIC. - LA ESPAÑOLA"),
    ("Harina 0000", 980, "$/kg", "HARINAS - FLORENCIA"),
    ("Huevos x6", 1_750, "$/pack", "HUEVOS FRESCOS"),
    ("Pan rallado", 3_130, "$/kg", "ALMACEN - PAN RALLADO NINO"),
    ("Aceite girasol 900ml", 2_750, "$/u", "1RA NECESIDAD - NATURA"),
    ("Pure tomate 520g", 785, "$/u", "ALMACEN - ARCOR"),
]
for insumo, precio, unidad, fuente in insumos_tabla:
    row += 1
    costo = round(precio * COST_FACTOR)
    wr(ws_rec, row, [insumo, precio, unidad, costo, "", "", "", fuente], ft=det_font)

# Tabla de costeo por receta
row += 2
seccion_row(ws_rec, row, "COSTEO POR RECETA", 8)
row += 1
write_header(ws_rec, row, ["Producto / Insumo", "Cantidad", "Unidad", "Precio Retail",
                           "Costo (x0.70)", "PVP", "Margen %", "% Ventas rotis."])

margen_ponderado = 0
for receta in recetas:
    row += 1
    wr(ws_rec, row, [receta["nombre"], "", "", "", "", receta["pvp"], "", f"{receta['pct_ventas']}%"], ft=bold, fl=lblue)
    total_costo = 0
    for item, qty, unit, precio_ret in receta["items"]:
        costo = round(qty * precio_ret * COST_FACTOR)
        total_costo += costo
        row += 1
        wr(ws_rec, row, [f"  {item}", round(qty, 2), unit, round(qty * precio_ret), costo, "", "", ""], ft=det_font)
    row += 1
    margen = (receta["pvp"] - total_costo) / receta["pvp"] * 100
    margen_ponderado += margen * receta["pct_ventas"] / 100
    fl_m = lgrn if margen > 25 else (lyel if margen > 0 else lred)
    wr(ws_rec, row, ["  COSTO TOTAL", "", "", "", total_costo, receta["pvp"],
                     f"{margen:.0f}%", ""], ft=bold, fl=fl_m)

row += 2
wr(ws_rec, row, ["MARGEN PONDERADO (estos productos)", "", "", "", "",
                 "", f"{margen_ponderado:.1f}%", ""], ft=bold, fl=lgrn)
row += 1
wr(ws_rec, row, ["Margen asumido en estructura de costos", "", "", "", "",
                 "", "30.0%", ""], ft=bold)
row += 1
nota_row(ws_rec, row, "  Factor costo 0.70 = precio retail x 70% (margen tipico supermercado ~30%)", 8)
row += 1
nota_row(ws_rec, row, "  Verde = margen >25% | Amarillo = margen 0-25% | Rojo = margen negativo", 8)


# =============================================
# HOJA 5: SENSIBILIZACION
# =============================================
ws4 = wb.create_sheet("Sensibilizacion")
set_widths(ws4, [22, 16, 14, 20, 16, 14])
titulo_row(ws4, 1, "ANALISIS DE SENSIBILIDAD - MARGEN DE CONTRIBUCION", 6)

ws4.cell(row=3, column=1, value="Escenarios variando Margen Bruto y Merma:").font = sub_font
write_header(ws4, 5, ["Escenario", "Margen Bruto %", "Merma %", "Margen Contrib ($)", "MC % s/Ventas", "Estado"])

escenarios = [
    ("Pesimista", 0.25, 0.08),
    ("Conservador", 0.28, 0.06),
    ("Base", 0.30, 0.05),
    ("Optimista", 0.33, 0.03),
    ("Ideal", 0.35, 0.02),
]

row = 6
for nom, mg, mr in escenarios:
    mb_e = round(V * mg)
    cv_e = comT + iibb_m + ich_m + round(V * mr)
    mc_e = mb_e - TF - cv_e
    pct_e = mc_e / V * 100
    est_e = "SUPERAVIT" if mc_e > 0 else "DEFICIT"
    fl_e = lgrn if mc_e > 0 else lred
    wr(ws4, row, [nom, f"{mg*100:.0f}%", f"{mr*100:.0f}%", mc_e, f"{pct_e:.1f}%", est_e], fl=fl_e)
    row += 1

row += 2
ws4.cell(row=row, column=1, value="Punto de equilibrio por escenario:").font = sub_font
row += 1
write_header(ws4, row, ["Escenario", "Margen %", "Merma %", "Vtas Equilibrio ($)", "Cobertura %", "Estado"])
row += 1
for nom, mg, mr in escenarios:
    vp_adj = (comT + iibb_m + ich_m + round(V * mr)) / V
    denom = mg - vp_adj
    if denom > 0:
        veq_e = round(TF / denom)
        cob_e = V / veq_e * 100
    else:
        veq_e = 0
        cob_e = 0
    fl_e = lgrn if cob_e > 100 else lred
    est_e = "OK" if cob_e > 100 else "INSUFICIENTE"
    wr(ws4, row, [nom, f"{mg*100:.0f}%", f"{mr*100:.0f}%", veq_e, f"{cob_e:.0f}%", est_e], fl=fl_e)
    row += 1

row += 2
nota_row(ws4, row, "  NOTA: Si obtiene certificado MiPyME, contribuciones SUSS bajan de 20.74% a ~18%, mejorando todos los escenarios.", 6)

# Guardar
out = "nino/ROTICERIA/Estructura_Costos_Rotiseria_NINO.xlsx"
wb.save(out)
print(f"Excel guardado: {out}")
print()
print("=== RESUMEN ===")
print(f"Ventas mensuales (desestac.): ${V:,}")
print(f"  (ventana: {WINDOW_DAYS}d, ${ventas_window:,.0f} bruto, idx={idx_weighted:.4f})")
print(f"CMV ({(1-MG)*100:.0f}%):                    -${CMV:,}")
print(f"Margen bruto ({MG*100:.0f}%):           ${MB:,}")
print(f"Costos fijos:                 -${TF:,}")
print(f"  Laboral:                    -${LAB:,}")
print(f"    Sueldos ({NE}x${SB_TOTAL:,}):  -${ST_total:,}")
print(f"    Contrib. (29.74% s/rem):   -${CP_m:,}")
print(f"    SAC mensualizado:         -${SAC:,}")
print(f"  Servicios:                  -${SERV:,}")
print(f"    Gas ({GAS_M3_ROTIS} m3):             -${gas_m:,}")
print(f"    Electricidad ({ELEC_KWH} kWh):   -${elec_m:,}")
print(f"  Otros fijos (estimados):    -${OTROS_F:,}")
print(f"Costos variables:             -${CV:,}")
print(f"RESULTADO:                    ${RES:,} ({RES/V*100:.1f}%)")
print(f"Pto equilibrio:               ${veq:,}/mes")
print(f"Cobertura:                    {cob:.0f}%")
