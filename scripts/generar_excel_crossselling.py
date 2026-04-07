"""Genera Excel de Cross-Selling - Rotiseria y Panaderia NINO
Analiza que otros productos compran los clientes en el mismo ticket."""
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import pandas as pd
import numpy as np
import warnings
warnings.filterwarnings("ignore")

# === ESTILOS ===
titulo_font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
sub_font = Font(name="Calibri", size=12, bold=True, color="1F4E79")
hdr_font = Font(name="Calibri", size=10, bold=True, color="FFFFFF")
norm = Font(name="Calibri", size=10)
bold = Font(name="Calibri", size=10, bold=True)
note_font = Font(name="Calibri", size=9, italic=True, color="996600")
det_font = Font(name="Calibri", size=9, color="666666")

tit_fill = PatternFill(start_color="1F4E79", end_color="1F4E79", fill_type="solid")
hdr_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
lblue = PatternFill(start_color="D6E4F0", end_color="D6E4F0", fill_type="solid")
lgrn = PatternFill(start_color="E2EFDA", end_color="E2EFDA", fill_type="solid")
lyel = PatternFill(start_color="FFF2CC", end_color="FFF2CC", fill_type="solid")
lred = PatternFill(start_color="FCE4EC", end_color="FCE4EC", fill_type="solid")
lgray = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
lorg = PatternFill(start_color="FDE9D9", end_color="FDE9D9", fill_type="solid")

brd = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin")
)


def write_header(ws, row, vals, c0=1):
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=c0 + i, value=v)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", wrap_text=True)
        c.border = brd


def wr(ws, row, vals, c0=1, ft=norm, fl=None):
    for i, v in enumerate(vals):
        c = ws.cell(row=row, column=c0 + i, value=v)
        c.font = ft
        c.border = brd
        if fl:
            c.fill = fl
        if isinstance(v, (int, float)) and not isinstance(v, bool):
            c.alignment = Alignment(horizontal="right")
            if isinstance(v, float):
                c.number_format = "#,##0"
            elif abs(v) >= 100:
                c.number_format = "#,##0"


def set_widths(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


def titulo_row(ws, row, text, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = titulo_font
    c.fill = tit_fill
    c.alignment = Alignment(horizontal="center")


def seccion_row(ws, row, text, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    c = ws.cell(row=row, column=1, value=text)
    c.font = sub_font
    c.fill = lblue


def nota_row(ws, row, text, cols=7):
    ws.merge_cells(start_row=row, start_column=1, end_row=row, end_column=cols)
    ws.cell(row=row, column=1, value=text).font = note_font


# === CARGAR DATOS ===
print("Cargando datos 2026...")
EXCEL_DIR = "H:/Mi unidad/PYMEINSIDE/nino/extraccion datos/2026/marzo"
excel_files = [
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes1.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes2.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes22.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes3.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes18-28.xlsx",
    f"{EXCEL_DIR}/ReporteMovimientosItemEnComprobantes31.xlsx",
]
dfs = []
for fp in excel_files:
    d = pd.read_excel(fp, header=3)
    dfs.append(d)
df = pd.concat(dfs, ignore_index=True)
df["Fecha"] = pd.to_datetime(df["Fecha"], format="%d/%m/%y", errors="coerce")
df = df[df["Fecha"].notna()]
df = df.drop_duplicates(subset=["Fecha", "Comprobante", "Código"])

# Ultimos 30 dias
date_max = df["Fecha"].max()
cutoff = date_max - pd.Timedelta(days=30)
df30 = df[df["Fecha"] >= cutoff].copy()
WINDOW_START = cutoff.strftime("%d/%m/%Y")
WINDOW_END = date_max.strftime("%d/%m/%Y")
DAYS = df30["Fecha"].dt.date.nunique()

print(f"Periodo: {WINDOW_START} - {WINDOW_END} ({DAYS} dias)")
print(f"Registros: {len(df30):,} | Tickets: {df30['Comprobante'].nunique():,}")


def analizar_crosssell(df30, depto_ancla, nombre_ancla):
    """Analiza cross-selling para un departamento ancla."""
    # Tickets que contienen el depto ancla
    tk_ancla = set(df30[df30["Departamento"] == depto_ancla]["Comprobante"].unique())
    n_tickets = len(tk_ancla)

    # Ventas del ancla en esos tickets
    ancla_data = df30[(df30["Comprobante"].isin(tk_ancla)) & (df30["Departamento"] == depto_ancla)]
    venta_ancla = ancla_data["Importe"].sum()

    # Cross-sell: otros deptos en esos tickets
    cross = df30[(df30["Comprobante"].isin(tk_ancla)) & (df30["Departamento"] != depto_ancla)]
    venta_cross = cross["Importe"].sum()

    # Por departamento
    by_dept = cross.groupby("Departamento").agg(
        venta=("Importe", "sum"),
        items=("Importe", "count"),
        tickets=("Comprobante", "nunique"),
    ).sort_values("venta", ascending=False)
    by_dept["pct_tickets"] = by_dept["tickets"] / n_tickets * 100
    by_dept["venta_prom_ticket"] = by_dept["venta"] / by_dept["tickets"]

    # Top productos cross-sell
    top_prods = cross.groupby(["Departamento", "Nombre"]).agg(
        venta=("Importe", "sum"),
        cantidad=("Cantidad", "sum"),
        tickets=("Comprobante", "nunique"),
    ).sort_values("venta", ascending=False).head(30)

    # Ticket promedio con y sin ancla
    tk_sin_ancla = set(df30["Comprobante"].unique()) - tk_ancla
    prom_con = df30[df30["Comprobante"].isin(tk_ancla)].groupby("Comprobante")["Importe"].sum().mean()
    prom_sin = df30[df30["Comprobante"].isin(tk_sin_ancla)].groupby("Comprobante")["Importe"].sum().mean() if tk_sin_ancla else 0

    return {
        "n_tickets": n_tickets,
        "venta_ancla": venta_ancla,
        "venta_cross": venta_cross,
        "venta_total": venta_ancla + venta_cross,
        "by_dept": by_dept,
        "top_prods": top_prods,
        "prom_con": prom_con,
        "prom_sin": prom_sin,
        "ratio": venta_cross / venta_ancla if venta_ancla > 0 else 0,
    }


print("Analizando cross-selling ROTISERIA...")
cs_rotis = analizar_crosssell(df30, "ROTISERIA", "Rotiseria")
print("Analizando cross-selling PANADERIA...")
cs_panad = analizar_crosssell(df30, "PANAD.ELAB.PROPIA", "Panaderia")

# === GENERAR EXCEL ===
wb = openpyxl.Workbook()


def escribir_hoja_crosssell(ws, nombre, cs, cols=7):
    """Escribe una hoja de cross-selling."""
    set_widths(ws, [35, 18, 12, 12, 14, 16, 16])

    titulo_row(ws, 1, f"CROSS-SELLING {nombre.upper()} - SUPERMERCADO NINO", cols)
    nota_row(ws, 2, f"  Periodo: {WINDOW_START} - {WINDOW_END} ({DAYS} dias) | Que compran los clientes junto con {nombre}?", cols)

    # --- RESUMEN ---
    r = 4
    seccion_row(ws, r, "RESUMEN", cols)
    r += 1
    write_header(ws, r, ["Concepto", "Valor", "", "", "", "", ""])
    r += 1
    wr(ws, r, [f"Tickets con {nombre}", cs["n_tickets"], "", "", "", "", ""], ft=bold)
    r += 1
    wr(ws, r, [f"Venta {nombre} en esos tickets", round(cs["venta_ancla"]), "", "", "", "", ""])
    r += 1
    wr(ws, r, ["Venta OTROS rubros en esos tickets", round(cs["venta_cross"]), "", "", "", "", ""], ft=bold, fl=lgrn)
    r += 1
    wr(ws, r, ["VENTA TOTAL en tickets con " + nombre, round(cs["venta_total"]), "", "", "", "", ""], ft=bold, fl=lyel)
    r += 1
    wr(ws, r, [f"Ratio cross-sell (otros / {nombre.lower()})", f"{cs['ratio']:.1f}x", "", "", "", "", f"Por cada $1 de {nombre.lower()}, se venden ${cs['ratio']:.1f} de otros rubros"], ft=bold, fl=lorg)
    r += 1
    wr(ws, r, ["", "", "", "", "", "", ""])
    r += 1
    wr(ws, r, [f"Ticket promedio CON {nombre.lower()}", round(cs["prom_con"]), "", "", "", "", ""], ft=bold)
    r += 1
    wr(ws, r, [f"Ticket promedio SIN {nombre.lower()}", round(cs["prom_sin"]), "", "", "", "", ""])
    r += 1
    diff = cs["prom_con"] - cs["prom_sin"]
    wr(ws, r, ["Diferencia", round(diff), "", "", "", "", f"El cliente de {nombre.lower()} gasta ${round(diff):,} mas por ticket"], ft=bold, fl=lgrn)

    # Mensualizado
    r += 2
    seccion_row(ws, r, "IMPACTO MENSUAL ESTIMADO", cols)
    r += 1
    mensual_factor = 30.4167 / DAYS
    venta_ancla_m = round(cs["venta_ancla"] * mensual_factor)
    venta_cross_m = round(cs["venta_cross"] * mensual_factor)
    write_header(ws, r, ["Concepto", "Mensual ($)", "", "", "", "", ""])
    r += 1
    wr(ws, r, [f"Venta {nombre} mensual", venta_ancla_m, "", "", "", "", ""])
    r += 1
    wr(ws, r, ["Cross-sell mensual (otros rubros)", venta_cross_m, "", "", "", "", ""], ft=bold, fl=lgrn)
    r += 1
    wr(ws, r, ["TOTAL impacto mensual", venta_ancla_m + venta_cross_m, "", "", "", "", ""], ft=bold, fl=lyel)
    r += 1
    nota_row(ws, r, f"  Si se cierra {nombre.lower()}, se perderian ~${venta_cross_m:,}/mes en ventas de otros rubros (cross-sell)", cols)

    # --- POR DEPARTAMENTO ---
    r += 2
    seccion_row(ws, r, f"CROSS-SELL POR DEPARTAMENTO (otros rubros comprados con {nombre.lower()})", cols)
    r += 1
    write_header(ws, r, ["Departamento", "Venta Total ($)", "Items", "Tickets", "% Tickets", "$/Ticket Prom", "Venta Mensual ($)"])
    r += 1
    for dept, row_data in cs["by_dept"].iterrows():
        venta_m = round(row_data["venta"] * mensual_factor)
        fl = lgrn if row_data["pct_tickets"] > 30 else (lyel if row_data["pct_tickets"] > 15 else None)
        wr(ws, r, [
            dept,
            round(row_data["venta"]),
            int(row_data["items"]),
            int(row_data["tickets"]),
            f"{row_data['pct_tickets']:.1f}%",
            round(row_data["venta_prom_ticket"]),
            venta_m,
        ], fl=fl)
        r += 1
    # Totales
    wr(ws, r, [
        "TOTAL CROSS-SELL",
        round(cs["venta_cross"]),
        int(cs["by_dept"]["items"].sum()),
        "",
        "",
        "",
        venta_cross_m,
    ], ft=bold, fl=lyel)
    r += 1
    nota_row(ws, r, "  Verde = presente en >30% tickets | Amarillo = 15-30% | Sin color = <15%", cols)

    # --- TOP PRODUCTOS ---
    r += 2
    seccion_row(ws, r, f"TOP 30 PRODUCTOS CROSS-SELL (comprados junto con {nombre.lower()})", cols)
    r += 1
    write_header(ws, r, ["Departamento", "Producto", "Venta ($)", "Cantidad", "Tickets", "$/Ticket", ""])
    r += 1
    for (dept, prod), row_data in cs["top_prods"].iterrows():
        wr(ws, r, [
            dept,
            prod[:40],
            round(row_data["venta"]),
            round(row_data["cantidad"]),
            int(row_data["tickets"]),
            round(row_data["venta"] / row_data["tickets"]),
            "",
        ])
        r += 1

    return r


# Hoja 1: Cross-sell Rotiseria
ws1 = wb.active
ws1.title = "CrossSell Rotiseria"
escribir_hoja_crosssell(ws1, "Rotiseria", cs_rotis)

# Hoja 2: Cross-sell Panaderia
ws2 = wb.create_sheet("CrossSell Panaderia")
escribir_hoja_crosssell(ws2, "Panaderia", cs_panad)

# Hoja 3: Resumen comparativo
ws3 = wb.create_sheet("Resumen")
set_widths(ws3, [35, 20, 20, 20])
titulo_row(ws3, 1, "RESUMEN CROSS-SELLING - ROTISERIA vs PANADERIA", 4)
nota_row(ws3, 2, f"  Periodo: {WINDOW_START} - {WINDOW_END} ({DAYS} dias)", 4)

r = 4
write_header(ws3, r, ["Concepto", "Rotiseria", "Panaderia", "Total"])
r += 1
items = [
    ("Tickets en el periodo", cs_rotis["n_tickets"], cs_panad["n_tickets"], ""),
    ("Venta propia ($)", round(cs_rotis["venta_ancla"]), round(cs_panad["venta_ancla"]), round(cs_rotis["venta_ancla"] + cs_panad["venta_ancla"])),
    ("Cross-sell ($)", round(cs_rotis["venta_cross"]), round(cs_panad["venta_cross"]), round(cs_rotis["venta_cross"] + cs_panad["venta_cross"])),
    ("Venta total tickets ($)", round(cs_rotis["venta_total"]), round(cs_panad["venta_total"]), round(cs_rotis["venta_total"] + cs_panad["venta_total"])),
    ("Ratio cross-sell", f"{cs_rotis['ratio']:.1f}x", f"{cs_panad['ratio']:.1f}x", ""),
    ("Ticket promedio CON ($)", round(cs_rotis["prom_con"]), round(cs_panad["prom_con"]), ""),
    ("Ticket promedio SIN ($)", round(cs_rotis["prom_sin"]), round(cs_panad["prom_sin"]), ""),
]
for concepto, v_rotis, v_panad, v_total in items:
    fl = lyel if "Ratio" in concepto or "total" in concepto.lower() else None
    ft = bold if "total" in concepto.lower() or "Ratio" in concepto or "Cross" in concepto else norm
    wr(ws3, r, [concepto, v_rotis, v_panad, v_total], ft=ft, fl=fl)
    r += 1

# Mensualizado
mensual_factor = 30.4167 / DAYS
r += 1
seccion_row(ws3, r, "IMPACTO MENSUAL ESTIMADO", 4)
r += 1
write_header(ws3, r, ["Concepto", "Rotiseria", "Panaderia", "Total"])
r += 1
items_m = [
    ("Venta propia mensual", round(cs_rotis["venta_ancla"] * mensual_factor), round(cs_panad["venta_ancla"] * mensual_factor)),
    ("Cross-sell mensual", round(cs_rotis["venta_cross"] * mensual_factor), round(cs_panad["venta_cross"] * mensual_factor)),
    ("TOTAL impacto mensual", round(cs_rotis["venta_total"] * mensual_factor), round(cs_panad["venta_total"] * mensual_factor)),
]
for concepto, v_r, v_p in items_m:
    fl = lyel if "TOTAL" in concepto else (lgrn if "Cross" in concepto else None)
    ft = bold if "TOTAL" in concepto or "Cross" in concepto else norm
    wr(ws3, r, [concepto, v_r, v_p, v_r + v_p], ft=ft, fl=fl)
    r += 1

r += 1
nota_row(ws3, r, "  CONCLUSION: Cerrar rotiseria/panaderia no solo pierde la venta propia, sino tambien el cross-sell de otros rubros", 4)

# Guardar
out = "nino/CrossSelling_NINO.xlsx"
wb.save(out)
print(f"\nExcel guardado: {out}")
print()
print("=== RESUMEN ===")
print(f"ROTISERIA: {cs_rotis['n_tickets']} tickets, venta propia ${cs_rotis['venta_ancla']:,.0f}, cross-sell ${cs_rotis['venta_cross']:,.0f} (ratio {cs_rotis['ratio']:.1f}x)")
print(f"PANADERIA: {cs_panad['n_tickets']} tickets, venta propia ${cs_panad['venta_ancla']:,.0f}, cross-sell ${cs_panad['venta_cross']:,.0f} (ratio {cs_panad['ratio']:.1f}x)")
print(f"Ticket promedio CON rotis: ${cs_rotis['prom_con']:,.0f} vs SIN: ${cs_rotis['prom_sin']:,.0f}")
print(f"Ticket promedio CON panad: ${cs_panad['prom_con']:,.0f} vs SIN: ${cs_panad['prom_sin']:,.0f}")
